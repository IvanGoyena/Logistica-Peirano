from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

URL_LOGIN = "https://app.digipwms.com"
URL_STOCK = "https://app.digipwms.com/Reportes/Stock"
USUARIO = "igoyena"
CLAVE = "0802"

CARPETA_DESCARGAS = Path(r"C:\Automatizacion_WMS\descargas\stock")
NOMBRE_TEMPORAL = "stock_detallado_temp"
NOMBRE_FINAL = "stock_detallado.xlsx"
NOMBRE_COLUMNA = "ContenedorNumero"
ESPERA_STOCK_DETALLADO_MS = 15000

def hacer_login(page):
    print("1. Abriendo login DIGIP...")
    page.goto(URL_LOGIN, wait_until="domcontentloaded")
    page.wait_for_load_state("networkidle")

    print("2. Completando usuario...")
    page.locator('input[name="username"], input[type="text"]').first.fill(USUARIO)

    print("3. Completando contraseña...")
    page.locator('input[name="password"], input[type="password"]').first.fill(CLAVE)

    print("4. Haciendo click en INGRESAR...")
    page.get_by_role("button", name="INGRESAR").click()

    print("5. Esperando ingreso al sistema...")
    page.wait_for_timeout(3000)
    page.wait_for_load_state("networkidle")

def leer_csv_tolerante(ruta_origen):
    df = pd.read_csv(
        ruta_origen,
        dtype=str,
        encoding="utf-8-sig",
        sep=",",
        engine="python",
        on_bad_lines="skip"
    )

    if len(df.columns) == 1 and ";" in str(df.columns[0]):
        df = pd.read_csv(
            ruta_origen,
            dtype=str,
            encoding="utf-8-sig",
            sep=";",
            engine="python",
            on_bad_lines="skip"
        )

    return df

def transformar_valor_contenedor(valor):
    if pd.isna(valor):
        return None

    texto = str(valor).strip()

    if texto == "":
        return None

    if any(c.isalpha() for c in texto):
        return texto

    texto = texto.replace(".", "").replace(",", ".")

    try:
        numero = float(texto)
        return int(round(numero, 0))
    except:
        return valor

def aplicar_formato_excel(ruta_destino, nombre_columna):
    wb = load_workbook(ruta_destino)
    ws = wb.active

    encabezados = {}
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=col).value
        if valor is not None:
            encabezados[str(valor).strip()] = col

    if nombre_columna not in encabezados:
        print(f"No se encontró la columna '{nombre_columna}' para aplicar formato Excel.")
        wb.close()
        return

    col_idx = encabezados[nombre_columna]

    for fila in range(2, ws.max_row + 1):
        celda = ws.cell(row=fila, column=col_idx)
        valor = celda.value

        if isinstance(valor, (int, float)) and valor is not None:
            celda.number_format = "0"

    wb.save(ruta_destino)
    wb.close()

    print(f"Formato Excel aplicado a la columna '{nombre_columna}'.")

def limpiar_y_generar_xlsx(ruta_origen, ruta_destino, nombre_columna):
    extension = ruta_origen.suffix.lower()

    try:
        print(f"Leyendo archivo descargado: {ruta_origen.name}")

        if extension == ".csv":
            df = leer_csv_tolerante(ruta_origen)
        elif extension in [".xlsx", ".xls"]:
            df = pd.read_excel(ruta_origen, dtype=str)
        else:
            print(f"Extensión no soportada para limpieza: {extension}")
            return

        if df.empty:
            print("El archivo quedó vacío después de leerlo.")
            return

        df.columns = [str(col).strip() for col in df.columns]

        if nombre_columna not in df.columns:
            print(f"No se encontró la columna '{nombre_columna}'.")
            print("Columnas detectadas:", list(df.columns))
            return

        df[nombre_columna] = df[nombre_columna].apply(transformar_valor_contenedor)

        with pd.ExcelWriter(ruta_destino, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)

        print(f"Archivo limpio generado: {ruta_destino}")

        aplicar_formato_excel(ruta_destino, nombre_columna)

    except Exception as e:
        print(f"Error al limpiar el archivo: {e}")

def main():
    CARPETA_DESCARGAS.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()

        try:
            hacer_login(page)

            print("6. Abriendo reporte de stock...")
            page.goto(URL_STOCK, wait_until="domcontentloaded")
            page.wait_for_load_state("networkidle")

            print("7. Click en Filtrar...")
            boton_filtrar = page.get_by_role("button", name="Filtrar")
            boton_filtrar.wait_for(state="visible", timeout=30000)
            boton_filtrar.click()
            page.wait_for_timeout(3000)

            print("8. Click en Stock Detallado...")
            tab_stock_detallado = page.get_by_text("Stock Detallado", exact=True)
            tab_stock_detallado.wait_for(state="visible", timeout=30000)
            tab_stock_detallado.click(force=True)

            print(f"9. Esperando {ESPERA_STOCK_DETALLADO_MS / 1000:.0f} segundos para que cargue el detalle...")
            page.wait_for_timeout(ESPERA_STOCK_DETALLADO_MS)

            print("10. Verificando que cargó la grilla de detalle...")
            page.get_by_role("columnheader", name="Ubicación").wait_for(state="visible", timeout=30000)
            page.get_by_role("columnheader", name="Código").wait_for(state="visible", timeout=30000)
            page.get_by_role("columnheader", name="Descripción").wait_for(state="visible", timeout=30000)

            print("11. Buscando botón Descargar completo...")
            boton_descargar = page.get_by_role("link", name="Descargar completo")
            boton_descargar.wait_for(state="visible", timeout=30000)

            print("12. Iniciando descarga...")
            with page.expect_download(timeout=60000) as download_info:
                boton_descargar.click()

            download = download_info.value
            nombre_sugerido = download.suggested_filename
            extension = Path(nombre_sugerido).suffix if nombre_sugerido else ".csv"

            ruta_temporal = CARPETA_DESCARGAS / f"{NOMBRE_TEMPORAL}{extension}"
            ruta_final = CARPETA_DESCARGAS / NOMBRE_FINAL

            download.save_as(str(ruta_temporal))
            print(f"Descarga temporal completada: {ruta_temporal}")

            print("13. Limpiando archivo y generando versión final para BI...")
            limpiar_y_generar_xlsx(ruta_temporal, ruta_final, NOMBRE_COLUMNA)

            print(f"Proceso finalizado. Archivo listo para BI: {ruta_final}")

        except PlaywrightTimeoutError as e:
            print("Timeout: no se encontró un elemento o no terminó la descarga.")
            print(e)
            try:
                page.screenshot(path="error_timeout.png", full_page=True)
                print("Se guardó captura: error_timeout.png")
            except:
                print("No se pudo guardar la captura porque la página ya estaba cerrada.")

        except Exception as e:
            print("Ocurrió un error durante la automatización.")
            print(e)
            try:
                page.screenshot(path="error_general.png", full_page=True)
                print("Se guardó captura: error_general.png")
            except:
                print("No se pudo guardar la captura porque la página ya estaba cerrada.")

        finally:
            try:
                context.close()
            except:
                pass
            try:
                browser.close()
            except:
                pass

if __name__ == "__main__":
    main()